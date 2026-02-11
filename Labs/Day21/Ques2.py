# 4. OpenPyXL / Pandas (Excel Handling)

# Question:
# You have an Excel file sales_data.xlsx with a sheet named "2025" that contains:

# Product	 Quantity	     Price
# A	         10	             50
# B	         20	             30
# C	         15	             40

# Use Pandas to read the Excel sheet into a DataFrame.

# Add a new column "Total" which is Quantity * Price.

# Save the updated DataFrame to a new Excel file called sales_summary.xlsx.

# (Bonus) Do the same using OpenPyXL without using Pandas






# Import required libraries
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook


# Using Pandas


# Read Excel file and sheet "2025"
df = pd.read_excel("sales_data.xlsx", sheet_name="2025")

# Add new column "Total"
df["Total"] = df["Quantity"] * df["Price"]

# Save updated DataFrame to new Excel file
df.to_excel("sales_summary.xlsx", index=False)

print("Pandas task completed. sales_summary.xlsx created.")



# Using OpenPyXL (Without Pandas)


# Load existing workbook and sheet
workbook = load_workbook("sales_data.xlsx")
sheet = workbook["2025"]

# Create a new workbook for output
new_workbook = Workbook()
new_sheet = new_workbook.active
new_sheet.title = "Summary"

# Write header row with new column
new_sheet.append(["Product", "Quantity", "Price", "Total"])

# Iterate through rows and calculate Total
for row in sheet.iter_rows(min_row=2, values_only=True):
    product, quantity, price = row
    total = quantity * price
    new_sheet.append([product, quantity, price, total])

# Save new Excel file
new_workbook.save("sales_summary_openpyxl.xlsx")

print("OpenPyXL task completed. sales_summary_openpyxl.xlsx created.")
